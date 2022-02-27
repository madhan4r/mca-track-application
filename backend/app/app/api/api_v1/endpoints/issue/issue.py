import os
import openpyxl
import pandas as pd
from pathlib import Path
from sqlalchemy.orm import Session
from typing import List
from fastapi import APIRouter, Body, Depends, HTTPException
from starlette.requests import Request
from starlette.responses import FileResponse

from app.crud.SearchFilter.db_model_search_filter import DbModelSearchFilter
from app.crud.issue.issue import issue
from app.api.deps import get_db, get_current_user
from app.models.issue.issue import IssueCreate, IssueUpdate, IssueSummary, IssueMinimal, IssueStatusCount, IssueProjectCount, IssueDetailCount
from app.db_models.user.user import User
from app.db_models.project.module import ProjectModule
from app.db_models.issue.issue import Issue
from app.db_models.issue.type import IssueType
from app.db_models.issue.status import IssueStatus
from app.db_models.issue.priority import IssuePriority
from app.db_models.project.project import Project

router = APIRouter()


@router.post('/', response_model=IssueSummary)
def create_issue(
    data_access_filter: IssueCreate,
    db_session: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return issue.create(db=db_session, obj_in=data_access_filter, created_by=current_user.user_id)


@router.get('/', response_model=List[IssueMinimal])
def get_issues(
    request: Request,
    db_session: Session = Depends(get_db),
    all_rows: bool = True,
    current_user: User = Depends(get_current_user),
) -> List[IssueMinimal]:
    return DbModelSearchFilter(
        db_session=db_session,
        request=request,
        db_model=Issue,
        join_db_models=[
            IssueType, IssueStatus, IssuePriority,
            ProjectModule, Project, User
        ],
        join_conditions=[
            Issue.type_id == IssueType.issue_type_id,
            Issue.status_id == IssueStatus.issue_status_id,
            Issue.priority_id == IssuePriority.issue_priority_id,
            Issue.module_id == ProjectModule.project_module_id,
            Issue.project_id == Project.project_id,
            Issue.created_by == User.user_id,
            Issue.assigned_to == User.user_id
        ],
        is_outer_join=True

    ).get_filtered_data(fetch_row_count=False, all_rows=all_rows)


@router.get('/count/', response_model=int)
def get_issues_count(
    request: Request,
    db_session: Session = Depends(get_db),
    fetch_row_count: bool = True,
    all_rows: bool = True,
    current_user: User = Depends(get_current_user),
) -> int:
    return DbModelSearchFilter(
        db_session=db_session,
        request=request,
        db_model=Issue,
        join_db_models=[
            IssueType, IssueStatus, IssuePriority,
            ProjectModule, Project, User
        ],
        join_conditions=[
            Issue.type_id == IssueType.issue_type_id,
            Issue.status_id == IssueStatus.issue_status_id,
            Issue.priority_id == IssuePriority.issue_priority_id,
            Issue.module_id == ProjectModule.project_module_id,
            Issue.project_id == Project.project_id,
            Issue.created_by == User.user_id,
            Issue.assigned_to == User.user_id
        ],
        is_outer_join=True

    ).get_filtered_data(fetch_row_count=fetch_row_count, all_rows=all_rows)


@router.get('/{issue_id}/get', response_model=IssueSummary)
def get_issue_by_id(
    issue_id: int,
    db_session: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[IssueSummary]:
    return issue.get(db=db_session, id=issue_id)


@router.put('/', response_model=IssueSummary)
def update_issue(
    data_access_filter: IssueUpdate,
    db_session: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    row = issue.get(db=db_session, id=data_access_filter.issue_id)
    if not row:
        raise HTTPException(status_code=400, detail='Requested issue does not exist')
    return issue.update(db=db_session, db_obj=row, obj_in=data_access_filter, created_by=current_user.user_id)


@router.delete("/{issue_id}")
def delete_issue(
    issue_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    issue_in = issue.get(db, id=issue_id)
    if not issue_in:
        raise HTTPException(
            status_code=404,
            detail="Requested issue does not exist in the system",
        )
    return issue.delete(db, id=issue_id)


@router.get('/count_by_status/', response_model=List[IssueStatusCount])
def get_issue_count_by_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[IssueStatusCount]:
    return issue.get_issue_count_by_status(db)


@router.get('/count_by_project/', response_model=List[IssueProjectCount])
def get_issue_count_by_project(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[IssueProjectCount]:
    return issue.get_issue_count_by_project(db)


@router.get('/count_by_detail/', response_model=List[IssueDetailCount])
def get_detailed_issues_count(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[IssueDetailCount]:
    return issue.get_detailed_issues_count(db)


@router.get("/export/excel")
def get_excel_data_for_issues(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    records = issue.get_issue_with_title(db)
    downloadable_file = f"track_issues_.xlsx"
    file_name = "file_store/"+downloadable_file

    df = pd.DataFrame(records)
    if os.path.isfile(file_name):
        print(f"delete file: {file_name}")
        os.remove(file_name)

    new_wb = openpyxl.Workbook()
    new_wb.save(filename=file_name)
    os.chmod(file_name, 0o777)

    SampleTemplate = Path("file_store/SampleTemplate.xlsx")

    wb1 = openpyxl.load_workbook(SampleTemplate)
    ws1 = wb1.worksheets[0]
    wb2 = openpyxl.load_workbook(file_name)
    ws2 = wb2.active

    mr = ws1.max_row
    mc = ws1.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = ws1.cell(row=i, column=j)

            ws2.cell(row=i, column=j).value = c.value

    wb2.save(filename=str(file_name))

    writer = pd.ExcelWriter(file_name, engine='openpyxl',
                            mode='a', if_sheet_exists='replace')

    writer.book = openpyxl.load_workbook(file_name)

    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

    reader = pd.read_excel(file_name)
    df.to_excel(writer, sheet_name="Sheet", index=False,
                header=True, startrow=len(reader))

    writer.close()

    return FileResponse(path=file_name, filename=downloadable_file,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
